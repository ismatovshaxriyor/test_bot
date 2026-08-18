"""Test natijalarini fayllarga eksport qilish"""
import os
import re
import tempfile
from typing import Dict
from database import Test

# XML/HTML da ruxsat etilmagan control belgilar (tab/newline'dan tashqari)
_CONTROL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")


def _clean_text(value) -> str:
    """Foydalanuvchi matnini tozalash: control belgilarni olib tashlash.

    (openpyxl Excel'ga, weasyprint HTML'ga ruxsatsiz control belgilar tushsa buziladi.)
    """
    return _CONTROL_RE.sub("", str(value or ""))




def get_grade(score: float) -> str:
    """Ballni daraja (grade) ga aylantirish"""
    if score >= 70:
        return "A+"
    elif score >= 65:
        return "A"
    elif score >= 60:
        return "B+"
    elif score >= 55:
        return "B"
    elif score >= 50:
        return "C+"
    elif score >= 46:
        return "C"
    else:
        return "-"

def export_to_excel(stats: Dict, test: Test) -> str:
    """Natijalarni Excel faylga yozish"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = f"Test {test.id}"

    # Daraja ranglari
    grade_fills = {
        'A+': PatternFill(start_color='22B14C', end_color='22B14C', fill_type='solid'),
        'A':  PatternFill(start_color='7BC67E', end_color='7BC67E', fill_type='solid'),
        'B+': PatternFill(start_color='FFF200', end_color='FFF200', fill_type='solid'),
        'B':  PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid'),
        'C+': PatternFill(start_color='F4B183', end_color='F4B183', fill_type='solid'),
        'C':  PatternFill(start_color='FF7F7F', end_color='FF7F7F', fill_type='solid'),
        '-':  PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
    }
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Sarlavha
    rasch_mode = test.scoring_mode == "rasch"
    ws.merge_cells('A1:H1' if rasch_mode else 'A1:E1')
    ws['A1'] = f"📊 Test natijasi — {test.id}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    # Ma'lumotlar
    ws['A3'] = "Test kodi:"
    ws['B3'] = test.id
    ws['A4'] = "Ishtirokchilar:"
    ws['B4'] = stats['total_submissions']
    ws['A5'] = "Savollar soni:"
    ws['B5'] = test.total_questions
    ws['A6'] = "Baholash:"
    ws['B6'] = "Rash modeli" if test.scoring_mode == "rasch" else "Oddiy"
    for r in range(3, 7):
        ws[f'A{r}'].font = Font(bold=True)

    # Jadval sarlavhalari
    row = 8

    if rasch_mode:
        headers = ["#", "Ism", "Umumiy ball", "Foiz", "Daraja", "1-Fan", "2-Fan"]
    else:
        headers = ["#", "Ism", "To'g'ri", "Jami", "Ball"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Ma'lumotlar
    submissions = stats['submissions']
    even_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    for i, sub in enumerate(submissions):
        num = i + 1
        r = row + 1 + i

        ball = round(sub.get('rasch_normalized', sub['percentage']) if rasch_mode else sub['percentage'], 2 if rasch_mode else 1)

        ws.cell(row=r, column=1, value=num).border = thin_border
        ws.cell(row=r, column=2, value=_clean_text(sub['user'])).border = thin_border

        if rasch_mode:
            grade = get_grade(ball)
            f1_val = sub.get('fan1_score', '-')
            f2_val = sub.get('fan2_score', '-')
            if grade == "-":
                f1_val = "-"
                f2_val = "-"

            ws.cell(row=r, column=3, value=ball).border = thin_border
            ws.cell(row=r, column=4, value=f"{int(round(sub['percentage']))}%").border = thin_border
            ws.cell(row=r, column=5, value=grade).border = thin_border
            ws.cell(row=r, column=6, value=f1_val).border = thin_border
            ws.cell(row=r, column=7, value=f2_val).border = thin_border

            fill = grade_fills.get(grade)
            if fill:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=r, column=c).fill = fill
        else:
            ws.cell(row=r, column=3, value=sub['correct']).border = thin_border
            ws.cell(row=r, column=4, value=sub['total']).border = thin_border
            ws.cell(row=r, column=5, value=ball).border = thin_border
            if i % 2 == 1:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=r, column=c).fill = even_fill

        # Markazlashtirish
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=2).alignment = Alignment(horizontal='left')

    # Ustun kengliklarini moslash
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12

    # ------------------ SAVOLLAR TAHLILI SHEET ------------------
    if rasch_mode and stats.get('question_stats'):
        ws2 = wb.create_sheet("Savollar tahlili")
        ws2.merge_cells('A1:D1')
        ws2['A1'] = "📋 Savollar tahlili"
        ws2['A1'].font = title_font
        ws2['A1'].alignment = Alignment(horizontal='center')

        rasch_data = stats.get('rasch', {})
        difficulties = rasch_data.get('question_difficulties', [])
        infits = rasch_data.get('question_infit', [])
        outfits = rasch_data.get('question_outfit', [])
        has_fit = bool(infits) and bool(outfits)

        q_headers = ["Savol #", "To'g'ri javoblar", "Foiz (%)", "Qiyinligi"]
        if has_fit:
            q_headers += ["Infit", "Outfit", "Mos kelish"]
        for col, header in enumerate(q_headers, 1):
            cell = ws2.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for i, qs in enumerate(stats['question_stats']):
            r = 4 + i
            ws2.cell(row=r, column=1, value=qs['index']).border = thin_border
            ws2.cell(row=r, column=2, value=qs['correct_count']).border = thin_border
            ws2.cell(row=r, column=3, value=qs['percentage']).border = thin_border

            if i < len(difficulties):
                diff = difficulties[i]
                if diff <= -1.5:
                    label = "Juda oson"
                elif diff <= -0.5:
                    label = "Oson"
                elif diff <= 0.5:
                    label = "O'rtacha"
                elif diff <= 1.5:
                    label = "Qiyin"
                else:
                    label = "Juda qiyin"
                ws2.cell(row=r, column=4, value=label).border = thin_border
            else:
                ws2.cell(row=r, column=4, value="-").border = thin_border

            if has_fit:
                infit = infits[i] if i < len(infits) else None
                outfit = outfits[i] if i < len(outfits) else None
                ws2.cell(row=r, column=5, value=infit if infit is not None else "-").border = thin_border
                ws2.cell(row=r, column=6, value=outfit if outfit is not None else "-").border = thin_border
                misfit = qs.get('misfit', False)
                ws2.cell(row=r, column=7, value="⚠️" if misfit else "✅").border = thin_border

            for c in range(1, len(q_headers) + 1):
                ws2.cell(row=r, column=c).alignment = Alignment(horizontal='center')

        ws2.column_dimensions['A'].width = 10
        ws2.column_dimensions['B'].width = 18
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 15
        if has_fit:
            ws2.column_dimensions['E'].width = 10
            ws2.column_dimensions['F'].width = 10
            ws2.column_dimensions['G'].width = 12

    # Faylni saqlash
    filepath = os.path.join(tempfile.gettempdir(), f"test_{test.id}.xlsx")
    wb.save(filepath)
    return filepath


def _register_pdf_font() -> str:
    """Unicode font (o'zbek harflari uchun) ro'yxatdan o'tkazish.

    Loyihadagi fonts/ papkasidan NotoSans ishlatiladi.
    Topilmasa — sistema fontlari (DejaVuSans, Arial Unicode) sinab ko'riladi.
    Oxirgi fallback: Helvetica (faqat ASCII).
    """
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    base_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fonts")

    candidates = [
        # Loyiha ichidagi fontlar (birinchi ustunlik)
        os.path.join(base_dir, "NotoSans-Regular.ttf"),
        # Linux (Docker)
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        # macOS
        "/Library/Fonts/Arial Unicode.ttf",
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
    ]
    for path in candidates:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont("UZ", path))
                return "UZ"
            except Exception:
                pass
    return "Helvetica"  # fallback — faqat ASCII ishlaydi


def _register_pdf_bold_font() -> str:
    """Bold font ro'yxatdan o'tkazish (sarlavhalar uchun)."""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    base_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fonts")

    candidates = [
        os.path.join(base_dir, "NotoSans-Bold.ttf"),
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
    ]
    for path in candidates:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont("UZ-Bold", path))
                return "UZ-Bold"
            except Exception:
                pass
    return "Helvetica-Bold"


def export_to_pdf(stats: Dict, test: Test) -> str:
    """Test natijalarini reportlab yordamida PDF formatda generatsiya qiladi.

    Args:
        stats: get_question_stats() dan qaytgan statistika dict.
        test: Test obyekti (test.id, test.total_questions, test.scoring_mode).

    Returns:
        str: yaratilgan PDF faylning to'liq yo'li.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )

    font = _register_pdf_font()
    bold_font = _register_pdf_bold_font()

    rasch_mode = test.scoring_mode == "rasch"
    mode_text = "Rash modeli" if rasch_mode else "Oddiy"

    # Chiqish faylini yaratish
    filepath = os.path.join(tempfile.gettempdir(), f"test_{test.id}.pdf")

    # PDF hujjatini sozlash
    doc = SimpleDocTemplate(
        filepath,
        pagesize=A4,
        title=f"Test {test.id}",
        topMargin=40,
        bottomMargin=30,
        leftMargin=35,
        rightMargin=35,
    )
    styles = getSampleStyleSheet()

    # Custom stillar
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontName=bold_font,
        fontSize=16,
        textColor=colors.HexColor("#1E293B"),
        alignment=TA_CENTER,
        spaceAfter=4,
    )
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontName=font,
        fontSize=10,
        textColor=colors.HexColor("#475569"),
        alignment=TA_CENTER,
        spaceAfter=14,
    )
    grade_note_style = ParagraphStyle(
        'GradeNote',
        parent=styles['Normal'],
        fontName=font,
        fontSize=8,
        textColor=colors.HexColor("#64748B"),
        spaceBefore=10,
    )

    # Sarlavha va umumiy ma'lumot
    elems = [
        Paragraph(f"Test #{test.id} — Natijalar", title_style),
        Paragraph(
            f"Ishtirokchilar: <b>{stats['total_submissions']} ta</b> &nbsp;·&nbsp; "
            f"Baholash: <b>{mode_text}</b>",
            subtitle_style,
        ),
        Spacer(1, 6),
    ]

    # Daraja ranglari (rasch mode uchun — foydalanuvchi rasmidagi shkala bo'yicha)
    grade_bg_colors = {
        'A+': colors.HexColor("#A8E6CF"),
        'A':  colors.HexColor("#C6EFCE"),
        'B+': colors.HexColor("#FFF200"),
        'B':  colors.HexColor("#FFD966"),
        'C+': colors.HexColor("#F4B183"),
        'C':  colors.HexColor("#FF7F7F"),
        '-':  colors.HexColor("#E0E0E0"),
    }

    # Jadval sarlavhalari (rasmdagi tartib bo'yicha)
    if rasch_mode:
        header_row = ["#", "Foydalanuvchi", "Umumiy ball", "Foiz", "Daraja", "1-Fan", "2-Fan"]
    else:
        header_row = ["#", "Foydalanuvchi", "To'g'ri", "Jami", "Ball"]
    rows = [header_row]

    # Natijalar satrlari
    for i, sub in enumerate(stats['submissions'], 1):
        name = _clean_text(sub['user']) or "—"
        if rasch_mode:
            ball_num = sub.get('rasch_normalized', sub['percentage'])
            grade = get_grade(ball_num)

            f1 = sub.get('fan1_score', '-')
            f2 = sub.get('fan2_score', '-')
            if grade == "-":
                f1_str = "-"
                f2_str = "-"
            else:
                f1_str = f"{f1:.1f}" if isinstance(f1, (int, float)) else str(f1)
                f2_str = f"{f2:.1f}" if isinstance(f2, (int, float)) else str(f2)

            row = [
                str(i),
                name,
                f"{ball_num:.2f}",
                f"{int(round(sub['percentage']))}%",
                grade,
                f1_str,
                f2_str,
            ]
        else:
            ball_num = sub['percentage']
            row = [str(i), name, str(sub['correct']), str(sub['total']), f"{ball_num:.1f}%"]
        rows.append(row)

    # Ustun kengliklari (A4 = 595pt, chap/o'ng margin 35pt -> avail = 525pt)
    avail_width = A4[0] - 70
    if rasch_mode:
        col_widths = [
            25,                    # #
            avail_width - 345,     # Foydalanuvchi
            75,                    # Umumiy ball
            55,                    # Foiz
            55,                    # Daraja
            67,                    # 1-Fan
            67,                    # 2-Fan
        ]
    else:
        col_widths = [
            30,
            avail_width - 240,
            70,
            70,
            70,
        ]

    table = Table(rows, repeatRows=1, hAlign="CENTER", colWidths=col_widths)

    # Asosiy jadval stili (rasmdagi kabi och ko'k header va qora chegaralar)
    style_commands = [
        ("FONTNAME",       (0, 0), (-1, 0),  bold_font),
        ("FONTNAME",       (0, 1), (-1, -1), font),
        ("FONTSIZE",       (0, 0), (-1, 0),  9),
        ("FONTSIZE",       (0, 1), (-1, -1), 9),
        ("BACKGROUND",     (0, 0), (-1, 0),  colors.HexColor("#B4C6E7")),
        ("TEXTCOLOR",      (0, 0), (-1, 0),  colors.black),
        ("GRID",           (0, 0), (-1, -1), 0.5, colors.black),
        ("ALIGN",          (0, 0), (-1, -1), "CENTER"),
        ("ALIGN",          (1, 0), (1, -1),  "LEFT"),
        ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",     (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING",  (0, 0), (-1, -1), 5),
        ("LEFTPADDING",    (0, 0), (-1, -1), 6),
        ("RIGHTPADDING",   (0, 0), (-1, -1), 6),
    ]

    # Qator ranglari
    if rasch_mode:
        for i, sub in enumerate(stats['submissions'], 1):
            ball_num = sub.get('rasch_normalized', sub['percentage'])
            grade = get_grade(ball_num)
            bg = grade_bg_colors.get(grade)
            if bg:
                style_commands.append(("BACKGROUND", (0, i), (-1, i), bg))
    else:
        style_commands.append(
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F4F4")])
        )

    table.setStyle(TableStyle(style_commands))
    elems.append(table)

    # Rasch daraja shkalasi izohi
    if rasch_mode:
        elems.append(Paragraph(
            "<b>Daraja shkalasi:</b> "
            "70+ → A+ · 65–69.9 → A · 60–64.9 → B+ · "
            "55–59.9 → B · 50–54.9 → C+ · 46–49.9 → C",
            grade_note_style,
        ))

    # PDF yaratish
    doc.build(elems)
    return filepath


def export_chart(stats: Dict, test: Test) -> str:
    """Natijalar grafikini yaratish"""
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    from matplotlib.patches import Patch

    q_stats = stats.get('question_stats', [])
    if not q_stats:
        return None

    total_q = len(q_stats)

    # Grafik o'lchami — savollar soniga qarab
    width = max(10, min(20, total_q * 0.45))
    fig, ax = plt.subplots(figsize=(width, 7))

    # Ma'lumotlar
    questions = [f"{qs['index']}" for qs in q_stats]
    percentages = [qs['percentage'] for qs in q_stats]

    # Ranglash
    colors = []
    for p in percentages:
        if p >= 80:
            colors.append('#27ae60')  # Yashil — oson
        elif p >= 60:
            colors.append('#f39c12')  # Sariq — o'rtacha
        elif p >= 40:
            colors.append('#e67e22')  # To'q sariq — qiyinroq
        else:
            colors.append('#c0392b')  # Qizil — qiyin

    # Fon zonalari
    ax.axhspan(80, 105, color='#27ae60', alpha=0.07)
    ax.axhspan(60, 80, color='#f39c12', alpha=0.07)
    ax.axhspan(40, 60, color='#e67e22', alpha=0.07)
    ax.axhspan(0, 40, color='#c0392b', alpha=0.07)

    # Zonalar chegarasi
    ax.axhline(y=80, color='#27ae60', linestyle='--', alpha=0.4, linewidth=1)
    ax.axhline(y=60, color='#f39c12', linestyle='--', alpha=0.4, linewidth=1)
    ax.axhline(y=40, color='#e67e22', linestyle='--', alpha=0.4, linewidth=1)

    # Zona nomlari (o'ng tomonda)
    ax.text(total_q - 0.5, 90, 'Oson', fontsize=9, color='#27ae60',
            fontweight='bold', ha='right', alpha=0.7)
    ax.text(total_q - 0.5, 70, "O'rtacha", fontsize=9, color='#f39c12',
            fontweight='bold', ha='right', alpha=0.7)
    ax.text(total_q - 0.5, 50, 'Qiyinroq', fontsize=9, color='#e67e22',
            fontweight='bold', ha='right', alpha=0.7)
    ax.text(total_q - 0.5, 20, 'Qiyin', fontsize=9, color='#c0392b',
            fontweight='bold', ha='right', alpha=0.7)

    # Ustunlar
    bars = ax.bar(range(total_q), percentages, color=colors,
                  edgecolor='white', linewidth=0.8, width=0.75,
                  zorder=3)

    # Foizni ustiga yozish
    for bar, pct in zip(bars, percentages):
        ax.text(bar.get_x() + bar.get_width()/2., bar.get_height() + 1.2,
                f'{pct:.0f}%', ha='center', va='bottom',
                fontsize=7 if total_q > 30 else 8, fontweight='bold', color='#333')

    # X o'qi
    ax.set_xticks(range(total_q))
    ax.set_xticklabels(questions, fontsize=7 if total_q > 30 else 9)
    if total_q > 25:
        plt.xticks(rotation=45, ha='right')

    ax.set_xlabel('Savol raqami', fontsize=12, fontweight='bold', labelpad=10)
    ax.set_ylabel("To'g'ri javoblar (%)", fontsize=12, fontweight='bold')
    ax.set_ylim(0, 108)
    ax.set_xlim(-0.5, total_q - 0.5)

    # Grid
    ax.yaxis.grid(True, alpha=0.2, linestyle='-')
    ax.set_axisbelow(True)

    # Ramka
    for spine in ['top', 'right']:
        ax.spines[spine].set_visible(False)

    # Sarlavha
    ax.set_title(f'Test {test.id} — Savollar qiyinligi tahlili',
                 fontsize=15, fontweight='bold', pad=15)

    # Legenda
    legend_elements = [
        Patch(facecolor='#27ae60', label='Oson (80%+)'),
        Patch(facecolor='#f39c12', label="O'rtacha (60-80%)"),
        Patch(facecolor='#e67e22', label='Qiyinroq (40-60%)'),
        Patch(facecolor='#c0392b', label='Qiyin (<40%)'),
    ]
    ax.legend(handles=legend_elements, loc='upper right', fontsize=9,
              framealpha=0.9, edgecolor='#ccc')

    # O'rtacha chiziq
    avg = sum(percentages) / len(percentages)
    ax.axhline(y=avg, color='#3498db', linestyle='-', alpha=0.6, linewidth=1.5, zorder=2)
    ax.text(0.5, avg + 1.5, f'O\'rtacha: {avg:.1f}%', fontsize=9,
            color='#3498db', fontweight='bold')

    plt.tight_layout()

    # Faylni saqlash
    filepath = os.path.join(tempfile.gettempdir(), f"chart_{test.id}.png")
    fig.savefig(filepath, dpi=150, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    return filepath


def export_grade_chart(stats: Dict, test: Test) -> str:
    """Ishtirokchilarning daraja (A+, A, B+, B, C+, C, NC) bo'yicha taqsimot grafigi.

    Daraja `get_grade()` bo'yicha aniqlanadi: Rash rejimida `rasch_normalized`
    balldan, oddiy rejimda foizdan (`percentage`) foydalaniladi.
    """
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt

    submissions = stats.get('submissions', [])
    if not submissions:
        return None

    order = ['A+', 'A', 'B+', 'B', 'C+', 'C', 'NC']
    bar_colors = {
        'A+': '#1e8449', 'A': '#27ae60', 'B+': '#f1c40f', 'B': '#f39c12',
        'C+': '#e67e22', 'C': '#d35400', 'NC': '#c0392b',
    }

    counts = {g: 0 for g in order}
    for sub in submissions:
        ball = sub.get('rasch_normalized', sub.get('percentage', 0))
        grade = get_grade(ball)
        counts['NC' if grade == '-' else grade] += 1

    total = len(submissions)
    values = [counts[g] for g in order]
    colors = [bar_colors[g] for g in order]

    fig, ax = plt.subplots(figsize=(9, 6.5))

    bars = ax.bar(order, values, color=colors, edgecolor='white',
                   linewidth=1, width=0.65, zorder=3)

    max_val = max(values) if values else 0
    for bar, val in zip(bars, values):
        if val == 0:
            continue
        pct = val / total * 100
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + max_val * 0.02,
                 f'{val}\n({pct:.0f}%)', ha='center', va='bottom',
                 fontsize=9, fontweight='bold', color='#333')

    ax.set_xlabel('Daraja', fontsize=12, fontweight='bold', labelpad=10)
    ax.set_ylabel('Foydalanuvchilar soni', fontsize=12, fontweight='bold')
    ax.set_ylim(0, max_val * 1.2 if max_val else 1)

    ax.set_title(f"Test {test.id} — Baholar bo'yicha taqsimot ({total} ishtirokchi)",
                 fontsize=14, fontweight='bold', pad=15)

    ax.yaxis.grid(True, alpha=0.2, linestyle='-')
    ax.set_axisbelow(True)
    for spine in ['top', 'right']:
        ax.spines[spine].set_visible(False)

    plt.tight_layout()

    filepath = os.path.join(tempfile.gettempdir(), f"gradechart_{test.id}.png")
    fig.savefig(filepath, dpi=150, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    return filepath
